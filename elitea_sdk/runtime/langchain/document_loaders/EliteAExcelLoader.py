# Copyright (c) 2026 EPAM Systems
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
import io
import logging
import os
import zipfile
from dataclasses import dataclass
from typing import Iterator, List, Optional, Union
import pandas as pd
from json import loads

from openpyxl import load_workbook
from xlrd import open_workbook
from langchain_core.documents import Document
from .EliteATableLoader import EliteATableLoader
from elitea_sdk.runtime.langchain.constants import LOADER_MAX_TOKENS_DEFAULT
from elitea_sdk.tools.utils.file_metadata import DEFAULT_MAX_OUTPUT_CHARS

logger = logging.getLogger(__name__)

cell_delimiter = " | "

EXCEL_MAX_REQUEST_ROWS = 10_000
EXCEL_MAX_IMAGE_COUNT = 32
EXCEL_SAMPLE_ROW_LIMIT = 10
EXCEL_MAX_FULL_READ_FILE_SIZE = 20 * 1024 * 1024
EXCEL_READ_LIMIT_ERROR = (
    "Excel read request exceeds safety limits. Use get_file_metadata to inspect sheets "
    "and call read_file with extra_params using a smaller start_row/end_row range."
)


@dataclass
class ExcelReadEstimate:
    """Structured estimate for an Excel read request."""

    sheets: List[dict]
    total_rows_workbook: int
    target_sheet: Optional[str]
    target_sheet_total_rows: int
    requested_start_row: int
    requested_end_row: int
    requested_rows: int
    sampled_rows: int
    sampled_chars: int
    estimated_output_chars: int
    embedded_images: int
    file_size_bytes: Optional[int]
    is_unbounded_read: bool
    violations: List[str]


class ExcelReadLimitExceeded(ValueError):
    """Raised on over-limit; carries the already-computed estimate."""

    def __init__(self, message: str, *, estimate: ExcelReadEstimate):
        super().__init__(message)
        self.estimate = estimate


# ---------------------------------------------------------------------------
# Lightweight, streaming-friendly helpers for sheet introspection and row-range
# reading. Added for EL-4389. Kept module-level (not on the loader class) so
# that the artifact / sharepoint / confluence wrappers can use them directly
# without going through the heavier chunking pipeline.
# ---------------------------------------------------------------------------


def _open_source(file_name: Optional[str], source: Union[str, bytes, io.BytesIO]):
    """Return (extension, openable) where openable is either a path string or a BytesIO."""
    if isinstance(source, (bytes, bytearray)):
        source = io.BytesIO(source)
    ref_name = file_name if file_name else (source if isinstance(source, str) else "")
    extension = os.path.splitext(ref_name)[-1].lower()
    return extension, source


def _read_binary_source(source: Union[str, bytes, io.BytesIO]) -> bytes:
    """Return workbook bytes for lightweight metadata inspection."""
    if isinstance(source, bytes):
        return source
    if isinstance(source, bytearray):
        return bytes(source)
    if isinstance(source, io.BytesIO):
        position = source.tell()
        try:
            source.seek(0)
            return source.read()
        finally:
            source.seek(position)
    with open(source, "rb") as file_obj:
        return file_obj.read()


def _get_source_size(source: Union[str, bytes, io.BytesIO]) -> Optional[int]:
    """Return raw workbook size in bytes when it can be determined cheaply."""
    if isinstance(source, bytes):
        return len(source)
    if isinstance(source, bytearray):
        return len(source)
    if isinstance(source, io.BytesIO):
        position = source.tell()
        try:
            source.seek(0, io.SEEK_END)
            return source.tell()
        finally:
            source.seek(position)
    try:
        return os.path.getsize(source)
    except OSError:
        return None


def _count_xlsx_images(source: Union[str, bytes, io.BytesIO],
                       file_name: Optional[str] = None) -> int:
    """Count embedded images in OOXML workbooks using zip metadata only."""
    extension, _ = _open_source(file_name, source)
    if extension not in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
        return 0

    if isinstance(source, str):
        archive_source = source
    else:
        archive_source = io.BytesIO(_read_binary_source(source))

    with zipfile.ZipFile(archive_source) as archive:
        return sum(
            1 for name in archive.namelist()
            if name.startswith("xl/media/") and not name.endswith("/")
        )


def _resolve_sheet_size(ws) -> tuple:
    """Return (max_row, max_column), falling back to a streaming scan when the
    worksheet has no <dimension> tag (e.g. written via openpyxl write_only=True),
    which leaves read_only max_row/max_column as None with no fallback in openpyxl.
    """
    if ws.max_row is not None:
        return ws.max_row, ws.max_column
    max_row = max_col = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    return max_row, max_col


def _list_sheets_and_sample(source: Union[str, bytes, io.BytesIO],
                            file_name: Optional[str] = None,
                            sample_limit: int = EXCEL_SAMPLE_ROW_LIMIT) -> tuple:
    """Open the workbook once and return sheet dims plus leading-row samples.

    Returns ``(sheets, sample_chars_by_name)`` where ``sheets`` matches
    ``list_excel_sheets`` and ``sample_chars_by_name`` maps each sheet to the
    formatted char count of its first ``sample_limit`` rows. A single open keeps
    the whole-workbook guard O(1) in sheet count instead of reopening per sheet.
    """
    extension, openable = _open_source(file_name, source)

    if extension in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
        wb = load_workbook(openable, read_only=True, data_only=True)
        try:
            sheets = []
            samples = {}
            for name in wb.sheetnames:
                ws = wb[name]
                max_row, max_col = _resolve_sheet_size(ws)
                sheets.append({
                    "name": name,
                    "max_row": max_row,
                    "max_column": max_col,
                })
                lines = [
                    _format_row(row) for row in
                    ws.iter_rows(min_row=1, max_row=sample_limit, values_only=True)
                ]
                samples[name] = len("\n".join(lines))
            return sheets, samples
        finally:
            wb.close()
    elif extension == '.xls':
        file_contents = None
        path = None
        if isinstance(openable, io.BytesIO):
            file_contents = openable.getvalue()
        else:
            path = openable
        wb = open_workbook(filename=path, file_contents=file_contents)
        sheets = []
        samples = {}
        for name in wb.sheet_names():
            sh = wb.sheet_by_name(name)
            sheets.append({"name": name, "max_row": sh.nrows, "max_column": sh.ncols})
            sample_end = min(sh.nrows, sample_limit)
            lines = [
                _format_row([sh.cell(r, c).value for c in range(sh.ncols)])
                for r in range(sample_end)
            ]
            samples[name] = len("\n".join(lines))
        return sheets, samples
    else:
        raise ValueError(f"Unsupported workbook extension for sheet listing: {extension}")


def check_excel_read_limits(source: Union[str, bytes, io.BytesIO],
                            *,
                            file_name: Optional[str] = None,
                            sheet_name: Optional[str] = None,
                            start_row: Optional[int] = None,
                            end_row: Optional[int] = None,
                            max_output_chars: int = DEFAULT_MAX_OUTPUT_CHARS,
                            raise_on_violation: bool = False) -> ExcelReadEstimate:
    """Estimate an Excel request and optionally reject unsafe reads."""
    extension, openable = _open_source(file_name, source)
    # A full workbook read materializes every sheet (see _read_xlsx), so for that
    # case list and sample leading rows in a single open — reopening per sheet is
    # ~quadratic in sheet count and defeats the guard's own purpose.
    maybe_whole_workbook = sheet_name is None and start_row is None and end_row is None
    sample_chars_by_name = None
    if maybe_whole_workbook:
        sheets, sample_chars_by_name = _list_sheets_and_sample(source, file_name=file_name)
    else:
        sheets = list_excel_sheets(source, file_name=file_name)
    total_rows_workbook = sum(sheet.get("max_row", 0) or 0 for sheet in sheets)

    target_sheet_info = None
    if sheet_name:
        target_sheet_info = next((sheet for sheet in sheets if sheet["name"] == sheet_name), None)
        if target_sheet_info is None:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {[sheet['name'] for sheet in sheets]}"
            )
    elif sheets:
        target_sheet_info = sheets[0]

    target_sheet = target_sheet_info["name"] if target_sheet_info else None
    target_sheet_total_rows = target_sheet_info.get("max_row", 0) if target_sheet_info else 0
    file_size_bytes = _get_source_size(source)
    whole_workbook_read = maybe_whole_workbook and bool(sheets)
    requested_start_row = start_row if start_row is not None and start_row > 0 else 1
    requested_end_row = end_row if end_row is not None else target_sheet_total_rows
    if target_sheet_total_rows > 0:
        requested_end_row = min(requested_end_row, target_sheet_total_rows)
    if whole_workbook_read:
        requested_rows = total_rows_workbook
    else:
        requested_rows = 0
        if target_sheet_total_rows > 0 and requested_start_row <= requested_end_row:
            requested_rows = requested_end_row - requested_start_row + 1

    embedded_images = 0
    if extension in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
        embedded_images = _count_xlsx_images(source, file_name=file_name)

    is_unbounded_read = start_row is None and end_row is None and (
        whole_workbook_read or requested_rows == target_sheet_total_rows
    )

    violations = []
    if is_unbounded_read and file_size_bytes is not None and file_size_bytes > EXCEL_MAX_FULL_READ_FILE_SIZE:
        violations.append(
            f"file size={file_size_bytes} exceeds full-read limit {EXCEL_MAX_FULL_READ_FILE_SIZE}"
        )
    if requested_rows > EXCEL_MAX_REQUEST_ROWS:
        violations.append(
            f"requested rows={requested_rows} exceeds limit {EXCEL_MAX_REQUEST_ROWS}"
        )
    if embedded_images > EXCEL_MAX_IMAGE_COUNT:
        violations.append(
            f"embedded images={embedded_images} exceeds limit {EXCEL_MAX_IMAGE_COUNT}"
        )

    sampled_rows = 0
    sampled_chars = 0
    estimated_output_chars = 0
    if not violations and whole_workbook_read:
        # Sum sampled/estimated output across every sheet the full read returns,
        # reusing the single-open samples from _list_sheets_and_sample above.
        for sheet in sheets:
            sheet_rows = sheet.get("max_row", 0) or 0
            if sheet_rows < 1:
                continue
            sampled_this_sheet = min(sheet_rows, EXCEL_SAMPLE_ROW_LIMIT)
            sheet_sampled = sample_chars_by_name.get(sheet["name"], 0)
            sampled_rows += sampled_this_sheet
            sampled_chars += sheet_sampled
            if sheet_rows > sampled_this_sheet and sampled_this_sheet > 0:
                estimated_output_chars += int(
                    (sheet_sampled / sampled_this_sheet) * sheet_rows
                )
            else:
                estimated_output_chars += sheet_sampled
    elif not violations and target_sheet and requested_rows > 0:
        sample_end_row = min(
            requested_end_row,
            requested_start_row + EXCEL_SAMPLE_ROW_LIMIT - 1,
        )
        sampled_rows = sample_end_row - requested_start_row + 1
        sample_result = read_excel_rows(
            openable,
            sheet_name=target_sheet,
            start_row=requested_start_row,
            end_row=sample_end_row,
            include_headers=False,
            file_name=file_name,
        )
        sampled_chars = len(sample_result.get("content", ""))
        estimated_output_chars = sampled_chars
        if requested_rows > sampled_rows and sampled_rows > 0:
            estimated_output_chars = int((sampled_chars / sampled_rows) * requested_rows)

    if sampled_chars > max_output_chars:
        violations.append(
            f"sampled output size={sampled_chars} exceeds limit {max_output_chars}"
        )
    elif estimated_output_chars > max_output_chars:
        violations.append(
            f"estimated output size={estimated_output_chars} exceeds limit {max_output_chars}"
        )

    estimate = ExcelReadEstimate(
        sheets=sheets,
        total_rows_workbook=total_rows_workbook,
        target_sheet=target_sheet,
        target_sheet_total_rows=target_sheet_total_rows,
        requested_start_row=requested_start_row,
        requested_end_row=requested_end_row,
        requested_rows=requested_rows,
        sampled_rows=sampled_rows,
        sampled_chars=sampled_chars,
        estimated_output_chars=estimated_output_chars,
        embedded_images=embedded_images,
        file_size_bytes=file_size_bytes,
        is_unbounded_read=is_unbounded_read,
        violations=violations,
    )

    if violations and raise_on_violation:
        raise ExcelReadLimitExceeded(
            f"{EXCEL_READ_LIMIT_ERROR} Details: {'; '.join(violations)}",
            estimate=estimate,
        )

    return estimate


def list_excel_sheets(source: Union[str, bytes, io.BytesIO],
                      file_name: Optional[str] = None) -> List[dict]:
    """List sheets in an Excel workbook without loading full content.

    Returns a list of dicts: ``{"name": str, "max_row": int, "max_column": int}``.
    Uses openpyxl ``read_only=True`` for ``.xlsx``/``.xlsm`` (streaming) and
    xlrd for legacy ``.xls``.
    """
    extension, openable = _open_source(file_name, source)

    if extension in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
        wb = load_workbook(openable, read_only=True, data_only=True)
        try:
            result = []
            for name in wb.sheetnames:
                ws = wb[name]
                max_row, max_col = _resolve_sheet_size(ws)
                result.append({
                    "name": name,
                    "max_row": max_row,
                    "max_column": max_col,
                })
            return result
        finally:
            wb.close()
    elif extension == '.xls':
        file_contents = None
        path = None
        if isinstance(openable, io.BytesIO):
            file_contents = openable.getvalue()
        else:
            path = openable
        wb = open_workbook(filename=path, file_contents=file_contents)
        result = []
        for name in wb.sheet_names():
            sh = wb.sheet_by_name(name)
            result.append({"name": name, "max_row": sh.nrows, "max_column": sh.ncols})
        return result
    else:
        raise ValueError(f"Unsupported workbook extension for sheet listing: {extension}")


def read_excel_rows(source: Union[str, bytes, io.BytesIO],
                    sheet_name: Optional[str] = None,
                    start_row: int = 1,
                    end_row: Optional[int] = None,
                    include_headers: bool = True,
                    header_row: int = 1,
                    file_name: Optional[str] = None) -> dict:
    """Read a row range from a single Excel sheet.

    Returns ``{"sheet_name": str, "start_row": int, "end_row": int,
    "total_rows": int, "content": str}`` where ``content`` is a pipe-delimited
    table (one row per line) optionally preceded by the header row.

    Uses streaming ``openpyxl.iter_rows`` for ``.xlsx``/``.xlsm`` so memory is
    O(range), not O(workbook).
    """
    if start_row is None or start_row < 1:
        start_row = 1
    if end_row is not None and end_row < start_row:
        raise ValueError(f"end_row ({end_row}) must be >= start_row ({start_row})")

    extension, openable = _open_source(file_name, source)

    if extension in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
        return _read_xlsx_rows(openable, sheet_name, start_row, end_row,
                               include_headers, header_row)
    elif extension == '.xls':
        return _read_xls_rows(openable, sheet_name, start_row, end_row,
                              include_headers, header_row)
    else:
        raise ValueError(f"Unsupported workbook extension for row reading: {extension}")


def _format_row(values) -> str:
    return cell_delimiter.join("" if v is None else str(v) for v in values)


def _format_row_cells(cells) -> str:
    """Format a row of Cell objects honoring hyperlinks (same as parse_sheet)."""
    parts = []
    for cell in cells:
        value = cell.value
        if cell.hyperlink:
            hyperlink = cell.hyperlink.target
            cell_value = value or ''
            parts.append(f"[{cell_value}]({hyperlink})")
        else:
            parts.append(str(value) if value is not None else "")
    return cell_delimiter.join(parts)


def _read_xlsx_rows(openable, sheet_name, start_row, end_row,
                    include_headers, header_row):
    wb = load_workbook(openable, read_only=True, data_only=True)
    try:
        if sheet_name is None:
            sheet_name = wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
        total_rows, _ = _resolve_sheet_size(ws)
        effective_end = min(end_row, total_rows) if end_row is not None else total_rows

        header_line = None
        if include_headers and header_row >= 1:
            for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
                header_line = _format_row(row)
                break

        body_rows = []
        if start_row <= effective_end:
            for row in ws.iter_rows(min_row=start_row, max_row=effective_end, values_only=True):
                body_rows.append(_format_row(row))

        lines = []
        if header_line is not None and (start_row > header_row or not body_rows
                                        or start_row != header_row):
            if not (start_row == header_row and body_rows
                    and body_rows[0] == header_line):
                lines.append(header_line)
        lines.extend(body_rows)

        return {
            "sheet_name": sheet_name,
            "start_row": start_row,
            "end_row": effective_end,
            "total_rows": total_rows,
            "content": "\n".join(lines),
        }
    finally:
        wb.close()


def _read_xls_rows(openable, sheet_name, start_row, end_row,
                   include_headers, header_row):
    file_contents = None
    path = None
    if isinstance(openable, io.BytesIO):
        file_contents = openable.getvalue()
    else:
        path = openable
    wb = open_workbook(filename=path, file_contents=file_contents)
    sheet_names = wb.sheet_names()
    if sheet_name is None:
        sheet_name = sheet_names[0]
    if sheet_name not in sheet_names:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available: {sheet_names}"
        )
    sh = wb.sheet_by_name(sheet_name)
    total_rows = sh.nrows
    effective_end = min(end_row, total_rows) if end_row is not None else total_rows

    def row_values(idx):
        return [sh.cell(idx, c).value for c in range(sh.ncols)]

    header_line = None
    if include_headers and 1 <= header_row <= total_rows:
        header_line = _format_row(row_values(header_row - 1))

    body_rows = []
    for r in range(start_row - 1, effective_end):
        body_rows.append(_format_row(row_values(r)))

    lines = []
    if header_line is not None and not (start_row == header_row and body_rows
                                        and body_rows[0] == header_line):
        lines.append(header_line)
    lines.extend(body_rows)

    return {
        "sheet_name": sheet_name,
        "start_row": start_row,
        "end_row": effective_end,
        "total_rows": total_rows,
        "content": "\n".join(lines),
    }

def build_excel_metadata_from_estimate(estimate: ExcelReadEstimate) -> dict:
    """Build the get_file_metadata-shaped dict directly from an estimate."""
    sheets = estimate.sheets
    read_limits = {
        "max_output_chars": DEFAULT_MAX_OUTPUT_CHARS,
        "full_read_allowed": estimate.is_unbounded_read and not estimate.violations,
        "excel_max_request_rows": EXCEL_MAX_REQUEST_ROWS,
        "excel_max_embedded_images": EXCEL_MAX_IMAGE_COUNT,
        "excel_full_read_max_bytes": EXCEL_MAX_FULL_READ_FILE_SIZE,
        "estimated_total_rows": estimate.total_rows_workbook,
        "estimated_request_rows": estimate.requested_rows,
        "estimated_output_chars": estimate.estimated_output_chars,
        "embedded_images": estimate.embedded_images,
        "file_size_bytes": estimate.file_size_bytes,
    }
    instruction = {
        "extra_params": {
            "sheet_name": (
                "string (optional) — name of the sheet to read. If omitted, "
                "the first sheet is used. Available sheets: "
                + ", ".join(s.get("name", "") for s in sheets) if sheets else
                "string (optional) — name of the sheet to read."
            ),
            "start_row": (
                "integer (1-indexed, inclusive) — first row to return. "
                "Defaults to 1 if omitted."
            ),
            "end_row": (
                "integer (1-indexed, inclusive) — last row to return. "
                "If omitted, reads to the end of the sheet."
            ),
            "include_headers": (
                "boolean (default true) — when true the header row is "
                "prepended to the output for column context."
            ),
            "header_row": (
                "integer (default 1) — 1-indexed row treated as the header."
            ),
            "evaluate_formulas": (
                "boolean (default false) — when true, formulas without "
                "cached values are evaluated using a Python engine. "
                "This can be slow for large workbooks."
            ),
        },
        "notes": (
            "For large workbooks, call read_file with a small "
            "start_row/end_row range to keep memory and tokens bounded. "
            "Pass extra_params as a JSON string. Full reads are rejected "
            "when workbook metadata suggests they exceed safe row, text, "
            "or embedded-image limits."
        ),
    }
    return {
        "unit": "rows",
        "total_rows": estimate.total_rows_workbook,
        "total_sheets": len(sheets),
        "sheets": sheets,
        "read_limits": read_limits,
        "instruction_for_readFile": instruction,
    }


class EliteAExcelLoader(EliteATableLoader):
    sheet_name: str = None
    file_name: str = None
    max_tokens: int = LOADER_MAX_TOKENS_DEFAULT
    add_header_to_chunks: bool = False
    header_row_number: int = 1
    evaluate_formulas: bool = False

    @classmethod
    def get_file_metadata(cls, *, filename: str,
                          file_content: Optional[bytes] = None,
                          file_size: Optional[int] = None) -> dict:
        """Return per-type metadata for Excel files (EL-4389).

        Called by ``tools.utils.file_metadata.get_file_metadata`` via the
        loader class from ``loaders_map``. Advertises available sheets and
        the extra_params that ``read_file`` can accept.
        """
        if not file_content:
            return build_excel_metadata_from_estimate(ExcelReadEstimate(
                sheets=[], total_rows_workbook=0, target_sheet=None,
                target_sheet_total_rows=0, requested_start_row=1,
                requested_end_row=0, requested_rows=0, sampled_rows=0,
                sampled_chars=0, estimated_output_chars=0, embedded_images=0,
                file_size_bytes=file_size, is_unbounded_read=False, violations=[],
            ))
        try:
            estimate = check_excel_read_limits(file_content, file_name=filename)
        except Exception as e:  # pylint: disable=broad-except
            logger.warning("Failed to list excel sheets for %s: %s", filename, e)
            estimate = ExcelReadEstimate(
                sheets=[], total_rows_workbook=0, target_sheet=None,
                target_sheet_total_rows=0, requested_start_row=1,
                requested_end_row=0, requested_rows=0, sampled_rows=0,
                sampled_chars=0, estimated_output_chars=0, embedded_images=0,
                file_size_bytes=file_size, is_unbounded_read=False, violations=[],
            )
        return build_excel_metadata_from_estimate(estimate)

    def __init__(self, **kwargs):
        if not kwargs.get('file_path'):
            file_content = kwargs.get('file_content')
            if file_content:
                self.file_name = kwargs.get('file_name')
                kwargs['file_path'] = io.BytesIO(file_content)
        else:
            self.file_name = kwargs.get('file_path')
        super().__init__(**kwargs)
        self.sheet_name = kwargs.get('sheet_name')
        # Row-range mode (EL-4389): when start_row/end_row supplied via
        # extra_params, get_content() returns a single pipe-delimited slice
        # instead of running the full chunking pipeline.
        self.start_row = kwargs.get('start_row')
        self.end_row = kwargs.get('end_row')
        self.include_headers = kwargs.get('include_headers', True)
        self.header_row = kwargs.get('header_row', 1)
        self.evaluate_formulas = bool(kwargs.get('evaluate_formulas', False))
        # Set and validate chunking parameters only once
        self.max_tokens = int(kwargs.get('max_tokens', LOADER_MAX_TOKENS_DEFAULT))
        self.add_header_to_chunks = bool(kwargs.get('add_header_to_chunks', False))
        header_row_number = kwargs.get('header_row_number', 1)
        # Validate header_row_number
        try:
            header_row_number = int(header_row_number)
            if header_row_number > 0:
                self.header_row_number = header_row_number
            else:
                self.header_row_number = 1
                self.add_header_to_chunks = False
        except (ValueError, TypeError):
            self.header_row_number = 1
            self.add_header_to_chunks = False

    def _is_row_range_mode(self) -> bool:
        return self.start_row is not None or self.end_row is not None

    def get_content(self):
        """Reads Excel file and returns dict of {sheet_name: list_of_chunks}.

        When ``start_row``/``end_row`` are set (EL-4389 row-range mode),
        returns a dict with the streaming row slice instead.
        """
        # Pass the caller's real start_row/end_row (preserving None) so the guard
        # can distinguish a genuine full read from a bounded range.
        check_excel_read_limits(
            self.file_path,
            file_name=self.file_name,
            sheet_name=self.sheet_name,
            start_row=self.start_row,
            end_row=self.end_row,
            raise_on_violation=True,
        )

        if self._is_row_range_mode():
            return read_excel_rows(
                self.file_path,
                sheet_name=self.sheet_name,
                start_row=self.start_row if self.start_row is not None else 1,
                end_row=self.end_row,
                include_headers=bool(self.include_headers),
                header_row=int(self.header_row) if self.header_row else 1,
                file_name=self.file_name,
            )

        # Determine file extension
        file_extension = os.path.splitext(self.file_name)[-1].lower()

        if file_extension == '.xlsx':
            return self._read_xlsx()
        elif file_extension == '.xls':
            return self._read_xls()
        else:
            raise ValueError(f"Unsupported file format: {file_extension}")

    def _compute_formula_values(self, file_path):
        """Uses formulas library to evaluate formula cells that openpyxl cannot resolve."""
        try:
            import formulas
            xl_model = formulas.ExcelModel().loads(file_path).finish()
            solution = xl_model.calculate()
            books = xl_model.write(solution=solution)
            computed = {}
            for book_data in books.values():
                for wb in book_data.values():
                    if not hasattr(wb, 'sheetnames'):
                        continue
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        for row in ws.iter_rows():
                            for cell in row:
                                if cell.value is not None:
                                    computed[(sheet_name, cell.coordinate)] = cell.value
            return computed
        except Exception as e:
            logger.debug("Formula evaluation skipped: %s", e)
            return {}

    def _read_xlsx(self):
        """Reads .xlsx files using openpyxl, with optional formula evaluation."""
        workbook = load_workbook(self.file_path, data_only=True)

        computed_values = {}
        if self.evaluate_formulas and isinstance(self.file_path, str):
            computed_values = self._compute_formula_values(self.file_path)

        sheets = workbook.sheetnames
        if self.sheet_name:
            if self.sheet_name in sheets:
                sheet_content = self.parse_sheet(workbook[self.sheet_name], computed_values)
            else:
                sheet_content = [f"Sheet '{self.sheet_name}' does not exist in the workbook."]
            return {self.sheet_name: sheet_content}
        else:
            return {name: self.parse_sheet(workbook[name], computed_values) for name in sheets}

    def _read_xls(self):
        """
        Reads .xls files using xlrd.
        """
        workbook = open_workbook(filename=self.file_name, file_contents=self.file_content)
        sheets = workbook.sheet_names()
        if self.sheet_name:
            if self.sheet_name in sheets:
                sheet = workbook.sheet_by_name(self.sheet_name)
                return {self.sheet_name: self.parse_sheet_xls(sheet)}
            else:
                return {self.sheet_name: [f"Sheet '{self.sheet_name}' does not exist in the workbook."]}
        else:
            # Dictionary comprehension for all sheets
            return {name: self.parse_sheet_xls(workbook.sheet_by_name(name)) for name in sheets}

    def parse_sheet(self, sheet, computed_values=None):
        """Parses a .xlsx sheet, extracting text/hyperlinks with formula evaluation fallback."""
        sheet_content = []

        for row in sheet.iter_rows():
            row_content = []
            for cell in row:
                value = cell.value
                # Fall back to computed formula value when openpyxl returns None
                # formulas library stores sheet names uppercased, so normalize for lookup
                if value is None and computed_values:
                    value = computed_values.get((sheet.title.upper(), cell.coordinate))

                if cell.hyperlink:
                    hyperlink = cell.hyperlink.target
                    cell_value = value or ''
                    row_content.append(f"[{cell_value}]({hyperlink})")
                else:
                    row_content.append(str(value) if value is not None else "")
            sheet_content.append(cell_delimiter.join(row_content))

        return self._format_sheet_content(sheet_content)

    def parse_sheet_xls(self, sheet):
        """
        Parses a single .xls sheet using xlrd, extracting text and hyperlinks, and formats them.
        """
        sheet_content = []

        # Extract hyperlink map (if available)
        hyperlink_map = getattr(sheet, 'hyperlink_map', {})

        for row_idx in range(sheet.nrows):
            row_content = []
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                cell_value = cell.value

                # Check if the cell has a hyperlink
                cell_address = (row_idx, col_idx)
                if cell_address in hyperlink_map:
                    hyperlink = hyperlink_map[cell_address].url_or_path
                    if cell_value:
                        row_content.append(f"[{cell_value}]({hyperlink})")
                else:
                    row_content.append(str(cell_value) if cell_value is not None else "")
            # Join the row content into a single line using `|` as the delimiter
            sheet_content.append(cell_delimiter.join(row_content))

        # Format the sheet content based on the return type
        return self._format_sheet_content(sheet_content)

    def _format_sheet_content(self, rows):
        """
        Specification:
        Formats a list of sheet rows into a list of string chunks according to the following rules:
        1. If max_tokens < 1, returns a single chunk (list of one string) with all rows joined by a newline ('\n').
           - If add_header_to_chunks is True and header_row_number is valid, the specified header row is prepended as the first line.
        2. If max_tokens >= 1:
           a. Each chunk is a string containing one or more rows, separated by newlines ('\n'), such that the total token count (as measured by tiktoken) does not exceed max_tokens.
           b. If add_header_to_chunks is True and header_row_number is valid, the specified header row is prepended once at the top of each chunk (not before every row).
           c. If a single row exceeds max_tokens, it is placed in its own chunk without splitting, with the header prepended if applicable.
        3. Returns: List[str], where each string is a chunk ready for further processing.
        """
        import tiktoken
        encoding = tiktoken.get_encoding('cl100k_base')

        # --- Inner functions ---
        def count_tokens(text):
            """Count tokens in text using tiktoken encoding."""
            return len(encoding.encode(text))

        def finalize_chunk(chunk_rows):
            """Join rows for a chunk, prepending header if needed."""
            if self.add_header_to_chunks and header:
                return '\n'.join([header] + chunk_rows)
            else:
                return '\n'.join(chunk_rows)
        # --- End inner functions ---

        # If max_tokens < 1, return all rows as a single chunk
        if self.max_tokens < 1:
            return ['\n'.join(rows)]

        # Extract header if needed
        header = None
        if self.add_header_to_chunks and rows:
            header_idx = self.header_row_number - 1
            header = rows.pop(header_idx)

        chunks = []  # List to store final chunks
        current_chunk = []  # Accumulate rows for the current chunk
        current_tokens = 0  # Token count for the current chunk

        for row in rows:
            row_tokens = count_tokens(row)
            # If row itself exceeds max_tokens, flush current chunk and add row as its own chunk (with header if needed)
            if row_tokens > self.max_tokens:
                if current_chunk:
                    chunks.append(finalize_chunk(current_chunk))
                    current_chunk = []
                    current_tokens = 0
                # Add the large row as its own chunk, with header if needed
                if self.add_header_to_chunks and header:
                    chunks.append(finalize_chunk([row]))
                else:
                    chunks.append(row)
                continue
            # If adding row would exceed max_tokens, flush current chunk and start new
            if current_tokens + row_tokens > self.max_tokens:
                if current_chunk:
                    chunks.append(finalize_chunk(current_chunk))
                current_chunk = [row]
                current_tokens = row_tokens
            else:
                current_chunk.append(row)
                current_tokens += row_tokens
        # Add any remaining rows as the last chunk
        if current_chunk:
            chunks.append(finalize_chunk(current_chunk))
        return chunks

    def load(self) -> list:
        """Loads Excel file into a list of Documents, one per chunk per sheet."""
        docs = []
        content_per_sheet = self.get_content()
        for sheet_name, content_chunks in content_per_sheet.items():
            metadata = {
                "source": f'{self.file_path}:{sheet_name}',
                "sheet_name": sheet_name,
                "file_type": "excel",
            }
            for chunk in content_chunks:
                docs.append(Document(page_content=chunk, metadata=metadata))
        return docs

    def read(self, lazy: bool = False):
        df = pd.read_excel(self.file_path, sheet_name=None, engine='calamine')
        docs = []
        for key in df.keys():
            if self.raw_content:
                docs.append(df[key].to_string())
            else:
                for record in loads(df[key].to_json(orient='records')):
                    docs.append(record)
        return docs

    def read_lazy(self) -> Iterator[dict]:
        df = pd.read_excel(self.file_path, sheet_name=None, engine='calamine')
        for key in df.keys():
            if self.raw_content:
                yield df[key].to_string()
            else:
                for record in loads(df[key].to_json(orient='records')):
                    yield record
        return
